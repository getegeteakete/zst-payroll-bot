"""
ポイント明細表 Excel を直接処理するモード

ユースケース:
  ZST社の事務員が、各乗務員シートに「運行行程」「荷姿」を手入力した
  ポイント明細表 Excel を LINE に送信
  → システムが追加ポイント・基本ポイント・集計を自動計算
  → 完成版 Excel と PDF を返却

CSV経由と違って:
  - 既存の入力データはそのまま尊重（上書きされない、空欄のセルだけ計算）
  - シート構造もそのまま維持
  - 月次override（無事故・愛車）のセル(Q53/S53)も既存値を尊重
"""
from openpyxl import load_workbook
from datetime import datetime, timedelta
from pathlib import Path
import subprocess
import tempfile
import shutil

from app.payroll_engine import (
    calc_route_points, calc_packing_bonus,
    cell_map_for_day,
)


# 月別所定労働日数
PRESCRIBED_DAYS = {
    (2024, 4): 22, (2024, 5): 21, (2024, 6): 20, (2024, 7): 23,
    (2024, 8): 22, (2024, 9): 21, (2024, 10): 23, (2024, 11): 20,
    (2024, 12): 22, (2025, 1): 22, (2025, 2): 19, (2025, 3): 21,
    (2025, 4): 22, (2025, 5): 21, (2025, 6): 21, (2025, 7): 22,
    (2025, 8): 21, (2025, 9): 22, (2025, 10): 22, (2025, 11): 20,
    (2025, 12): 22, (2026, 1): 22, (2026, 2): 20, (2026, 3): 21,
    (2026, 4): 22, (2026, 5): 22,
}


def detect_target_year_month(ws):
    """A4セルから対象年月を取得"""
    a4 = ws["A4"].value
    if isinstance(a4, datetime):
        return (a4.year, a4.month)
    if isinstance(a4, (int, float)):
        d = datetime(1899, 12, 30) + timedelta(days=int(a4))
        return (d.year, d.month)
    return None


def process_one_sheet(ws, year, month, prescribed_days):
    """
    1シート（1乗務員分）を処理
    - 既に入力済みの運行行程・荷姿から追加ポイントを計算
    - 基本ポイントを設定
    - 既存のH/S列の数値はチェック（不整合は補正）
    
    Returns: 処理サマリ
    """
    days_in_month = (
        datetime(year + (1 if month == 12 else 0),
                 (1 if month == 12 else month + 1), 1)
        - timedelta(days=1)
    ).day
    
    work_days = 0
    routes_processed = 0
    routes_unknown = 0
    routes_corrected = 0  # 既存値を上書きした件数
    routes_filled = 0     # 空欄に新規記入した件数
    
    for day in range(1, days_in_month + 1):
        if day <= 16:
            cmap = cell_map_for_day(day - 1, "left")
        else:
            cmap = cell_map_for_day(day - 17, "right")
        
        # その日の運行行程を見て、稼働日かどうか判定
        is_working = False
        for i in range(3):
            col_r, row_r = cmap["route"][i]
            route = ws[f"{col_r}{row_r}"].value
            if route:
                route_s = str(route).strip()
                if route_s and "休み" not in route_s and "特別休暇" not in route_s and "代休" not in route_s:
                    is_working = True
                    break
        
        if is_working:
            work_days += 1
            # 基本ポイント9000を設定（既存値があれば上書きしない）
            col_b, row_b = cmap["base_pt"]
            existing_base = ws[f"{col_b}{row_b}"].value
            if existing_base is None or existing_base == 0:
                ws[f"{col_b}{row_b}"] = 9000
        
        # 各便ごとに追加ポイント計算
        for i in range(3):
            col_r, row_r = cmap["route"][i]
            col_p, row_p = cmap["packing"][i]
            col_a, row_a = cmap["add_pt"][i]
            
            route = ws[f"{col_r}{row_r}"].value
            packing = ws[f"{col_p}{row_p}"].value
            existing_add_pt = ws[f"{col_a}{row_a}"].value
            
            if not route:
                continue
            route_s = str(route).strip()
            if not route_s or "休み" in route_s or "特別休暇" in route_s or "代休" in route_s:
                continue
            
            # 推定値を計算
            rt_pt, _, conf = calc_route_points(route_s, str(packing or "").strip())
            pkg_pt = calc_packing_bonus(str(packing or "").strip())
            
            if rt_pt is not None:
                estimated = rt_pt + pkg_pt
                routes_processed += 1
                
                if existing_add_pt is None or existing_add_pt == 0:
                    # 空欄 → 推定値を記入
                    ws[f"{col_a}{row_a}"] = estimated
                    routes_filled += 1
                elif isinstance(existing_add_pt, (int, float)) and existing_add_pt != estimated:
                    # 既存値あり、推定値と違う → 既存値を尊重（事務員の入力優先）
                    pass
            else:
                routes_unknown += 1
    
    return {
        "work_days": work_days,
        "routes_processed": routes_processed,
        "routes_unknown": routes_unknown,
        "routes_filled": routes_filled,
        "routes_corrected": routes_corrected,
        "prescribed_days": prescribed_days,
    }


def process_payroll_xlsx(input_xlsx_path: Path, output_dir: Path = None) -> dict:
    """
    メインエントリ：受信したxlsxを直接処理
    
    Returns:
        {
            "results": [{乗務員別サマリ}, ...],
            "completed_xlsx_path": Path,
            "completed_pdf_path": Path,
        }
    """
    output_dir = output_dir or Path(tempfile.mkdtemp(prefix="payroll_xlsx_"))
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # コピーして編集（ファイル名から空白を除去：URLリンク化対策）
    safe_input_name = input_xlsx_path.name.replace(" ", "_").replace("　", "_")
    output_xlsx = output_dir / f"完成_{safe_input_name}"
    shutil.copy(input_xlsx_path, output_xlsx)
    
    wb = load_workbook(output_xlsx)
    
    results = []
    target_ym = None
    
    for sn in wb.sheetnames:
        if sn in ("明細原本", "Sheet1", "Sheet2"):
            continue
        ws = wb[sn]
        
        # 対象年月をシートから取得
        if target_ym is None:
            target_ym = detect_target_year_month(ws)
            if target_ym is None:
                raise ValueError("A4セルから年月を取得できません。テンプレ形式を確認してください")
        
        year, month = target_ym
        prescribed_days = PRESCRIBED_DAYS.get((year, month), 22)
        
        summary = process_one_sheet(ws, year, month, prescribed_days)
        
        # メタ情報を更新
        # 乗務員名（L53が空ならシート名から）
        if not ws["L53"].value:
            ws["L53"] = sn
        
        # 配車貢献（既存値があれば尊重、空欄なら所定日数達成判定）
        if ws["U50"].value is None:
            if summary["work_days"] >= prescribed_days:
                ws["U50"] = 10000
            else:
                ws["U50"] = 0
        
        # 出勤加算（既存値・既存数式があれば尊重、空欄なら標準数式を入れる）
        if ws["N50"].value is None:
            ws["N50"] = f"=MAX((L50-{prescribed_days})*5000,0)"
        
        # 愛車・無事故が空欄なら0（既存値があれば尊重）
        if ws["Q53"].value is None:
            ws["Q53"] = 0
        if ws["S53"].value is None:
            ws["S53"] = 0
        
        results.append({
            "driver_name": ws["L53"].value or sn,
            "sheet_name": sn,
            **summary,
        })
    
    wb.save(output_xlsx)
    
    # 数式再計算 → PDF変換
    recalc_excel(output_xlsx)
    
    # 集計値を再取得
    wb_recalc = load_workbook(output_xlsx, data_only=True)
    for r in results:
        ws = wb_recalc[r["sheet_name"]]
        r["summary"] = {
            "稼働日数": ws["L50"].value,
            "出勤加算ポイント": ws["N50"].value,
            "基本ポイント合計": ws["O50"].value,
            "追加ポイント合計": ws["S50"].value,
            "配車貢献ポイント": ws["U50"].value,
            "愛車ポイント": ws["Q53"].value,
            "無事故ポイント": ws["S53"].value,
            "ポイント合計": ws["U53"].value,
        }
    
    # PDF変換
    pdf_path = output_dir / output_xlsx.with_suffix(".pdf").name
    convert_to_pdf(output_xlsx, pdf_path)
    
    return {
        "results": results,
        "completed_xlsx_path": output_xlsx,
        "completed_pdf_path": pdf_path,
        "year_month": target_ym,
    }


def recalc_excel(xlsx_path: Path):
    """LibreOfficeで開いて保存し直すことで数式を再計算"""
    tmpdir = tempfile.mkdtemp()
    subprocess.run([
        "libreoffice", "--headless", "--calc",
        "--convert-to", "xlsx",
        "--outdir", tmpdir,
        str(xlsx_path)
    ], check=True, timeout=60)
    converted = Path(tmpdir) / xlsx_path.name
    if converted.exists():
        converted.replace(xlsx_path)


def convert_to_pdf(xlsx_path: Path, pdf_path: Path):
    """PDF変換"""
    tmpdir = tempfile.mkdtemp()
    subprocess.run([
        "libreoffice", "--headless",
        "--convert-to", "pdf",
        "--outdir", tmpdir,
        str(xlsx_path)
    ], check=True, timeout=60)
    converted = Path(tmpdir) / (xlsx_path.stem + ".pdf")
    if converted.exists():
        converted.replace(pdf_path)

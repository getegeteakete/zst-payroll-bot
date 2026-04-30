"""
給与計算オーケストレーター
配車CSV → ポイント明細表Excel + PDF 生成
"""
from openpyxl import load_workbook
from datetime import datetime, timedelta
from collections import defaultdict
import csv
import io
import subprocess
import os
import tempfile
from pathlib import Path

from app.payroll_engine import (
    calc_route_points, calc_packing_bonus,
    cell_map_for_day, date_to_serial,
)

# 月別所定労働日数（2025年6月〜2026年5月）
PRESCRIBED_DAYS_2025 = {
    (2025, 6): 21, (2025, 7): 22, (2025, 8): 21, (2025, 9): 22,
    (2025, 10): 22, (2025, 11): 20, (2025, 12): 22,
    (2026, 1): 22, (2026, 2): 20, (2026, 3): 21,
    (2026, 4): 22, (2026, 5): 22,
}

TEMPLATE_PATH = Path(__file__).parent.parent / "templates" / "blank_template.xlsx"


def parse_dispatch_csv(csv_text: str) -> tuple[dict, dict]:
    """
    配車CSVをパース。柔軟にカラム名を解釈する。
    必須: 日付、便、運行行程、荷姿
    オプション: 乗務員名（あれば乗務員別に分割）
    
    返り値: (driver_name, dispatch_by_date), warnings
    """
    f = io.StringIO(csv_text)
    reader = csv.DictReader(f)
    fieldnames = reader.fieldnames or []
    warnings = []

    # カラム名のゆらぎを正規化
    col_map = {}
    for col in fieldnames:
        col_lower = col.strip().lower()
        if col in ("日付", "date") or "日付" in col:
            col_map["date"] = col
        elif col in ("便", "便号", "trip"):
            col_map["trip"] = col
        elif "運行" in col or "ルート" in col or "route" in col_lower:
            col_map["route"] = col
        elif "荷姿" in col or "荷物" in col or "package" in col_lower:
            col_map["packing"] = col
        elif "乗務員" in col or "ドライバー" in col or "名前" in col or "driver" in col_lower:
            col_map["driver"] = col

    for required in ("date", "route"):
        if required not in col_map:
            raise ValueError(f"CSVに必須カラムがありません: {required}")

    driver_name = "（名称未指定）"
    by_date = defaultdict(lambda: [None, None, None])

    for row in reader:
        d_raw = row.get(col_map["date"], "").strip()
        if not d_raw:
            continue
        # 日付パース
        date_str = _normalize_date(d_raw)
        if not date_str:
            warnings.append(f"日付不正: {d_raw}")
            continue

        trip = row.get(col_map.get("trip", ""), "①").strip() or "①"
        trip_idx = {"①": 0, "②": 1, "③": 2, "1": 0, "2": 1, "3": 2}.get(trip, 0)

        route = row.get(col_map["route"], "").strip()
        packing = row.get(col_map.get("packing", ""), "").strip() if "packing" in col_map else ""

        if "driver" in col_map:
            driver_name = row.get(col_map["driver"], driver_name).strip() or driver_name

        by_date[date_str][trip_idx] = {
            "運行行程": route,
            "荷姿": packing,
        }

    return driver_name, dict(by_date), warnings


def _normalize_date(d_raw: str) -> str | None:
    """日付文字列を YYYY-MM-DD に正規化"""
    d_raw = d_raw.strip()
    formats = ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"]
    for fmt in formats:
        try:
            return datetime.strptime(d_raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Excelシリアル値
    try:
        n = int(float(d_raw))
        if 40000 < n < 60000:
            return (datetime(1899, 12, 30) + timedelta(days=n)).strftime("%Y-%m-%d")
    except ValueError:
        pass
    return None


def populate_template(
    output_path: Path,
    sheet_name: str,
    dispatch_by_date: dict,
    driver_name: str,
    target_year_month: tuple[int, int],
    prescribed_days: int,
):
    """テンプレに流し込み"""
    wb = load_workbook(TEMPLATE_PATH)
    if sheet_name not in wb.sheetnames:
        src = wb["明細原本"]
        ws = wb.copy_worksheet(src)
        ws.title = sheet_name
    else:
        ws = wb[sheet_name]

    year, month = target_year_month
    first_day = datetime(year, month, 1)
    if month == 12:
        last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = datetime(year, month + 1, 1) - timedelta(days=1)
    days_in_month = last_day.day

    ws["A4"] = date_to_serial(first_day)

    work_days = 0
    fill_log = []

    for day in range(1, days_in_month + 1):
        date_obj = datetime(year, month, day)
        date_str = date_obj.strftime("%Y-%m-%d")
        trips = dispatch_by_date.get(date_str, [None, None, None])

        if day <= 16:
            cmap = cell_map_for_day(day - 1, "left")
        else:
            cmap = cell_map_for_day(day - 17, "right")

        is_working = any(
            trip and trip["運行行程"] and "休み" not in trip["運行行程"]
            and "特別休暇" not in trip["運行行程"]
            for trip in trips
        )
        if is_working:
            work_days += 1
            col, row = cmap["base_pt"]
            ws[f"{col}{row}"] = 9000

        for i, trip in enumerate(trips):
            if not trip:
                continue
            route = trip["運行行程"]
            packing = trip["荷姿"]
            col_r, row_r = cmap["route"][i]
            col_p, row_p = cmap["packing"][i]
            ws[f"{col_r}{row_r}"] = route or ""
            ws[f"{col_p}{row_p}"] = packing or ""

            if route and "休み" not in str(route) and "特別休暇" not in str(route):
                route_pt, cat, conf = calc_route_points(route, packing)
                packing_pt = calc_packing_bonus(packing)
                if route_pt is not None:
                    add_pt = route_pt + packing_pt
                    col_a, row_a = cmap["add_pt"][i]
                    ws[f"{col_a}{row_a}"] = add_pt
                    fill_log.append({"日付": date_str, "便": ["①","②","③"][i],
                                    "行程": route, "推定": add_pt, "信頼度": conf})
                else:
                    fill_log.append({"日付": date_str, "便": ["①","②","③"][i],
                                    "行程": route, "推定": None, "信頼度": "low"})

    # メタデータ
    ws["L53"] = driver_name
    if ws["Q53"].value is None or not isinstance(ws["Q53"].value, (int, float)):
        ws["Q53"] = 0
    if ws["S53"].value is None or not isinstance(ws["S53"].value, (int, float)):
        ws["S53"] = 0
    ws["N50"] = f"=MAX((L50-{prescribed_days})*5000,0)"

    if work_days < prescribed_days:
        ws["U50"] = 0
    else:
        ws["U50"] = 10000

    wb.save(output_path)
    return work_days, fill_log


def recalc_excel(xlsx_path: Path):
    """LibreOfficeで開いて保存しなおすことで数式を再計算"""
    tmpdir = tempfile.mkdtemp()
    subprocess.run([
        "libreoffice", "--headless", "--calc",
        "--convert-to", "xlsx",
        "--outdir", tmpdir,
        str(xlsx_path)
    ], check=True, timeout=60)
    converted = Path(tmpdir) / xlsx_path.name
    if converted.exists():
        converted.replace(xlsx_path)


def convert_to_pdf(xlsx_path: Path, pdf_path: Path):
    """ExcelをPDFに変換"""
    tmpdir = tempfile.mkdtemp()
    subprocess.run([
        "libreoffice", "--headless",
        "--convert-to", "pdf",
        "--outdir", tmpdir,
        str(xlsx_path)
    ], check=True, timeout=60)
    converted = Path(tmpdir) / (xlsx_path.stem + ".pdf")
    if converted.exists():
        converted.replace(pdf_path)


def process_payroll(
    csv_text: str,
    target_year_month: tuple[int, int] | None = None,
    output_dir: Path | None = None,
) -> dict:
    """
    メインエントリポイント
    
    Returns:
        {
            "xlsx_path": Path,
            "pdf_path": Path,
            "driver_name": str,
            "work_days": int,
            "summary": dict,
            "warnings": [str],
            "low_conf_count": int,
        }
    """
    output_dir = output_dir or Path(tempfile.mkdtemp())
    output_dir.mkdir(parents=True, exist_ok=True)

    driver_name, dispatch, warnings = parse_dispatch_csv(csv_text)

    # 対象年月を推定（CSVの最頻年月）
    if target_year_month is None:
        from collections import Counter
        ym_counter = Counter(
            (datetime.strptime(d, "%Y-%m-%d").year, datetime.strptime(d, "%Y-%m-%d").month)
            for d in dispatch.keys()
        )
        if not ym_counter:
            raise ValueError("CSVに有効な日付データがありません")
        target_year_month = ym_counter.most_common(1)[0][0]

    prescribed_days = PRESCRIBED_DAYS_2025.get(target_year_month, 22)

    safe_name = driver_name.replace("（", "_").replace("）", "_").replace(" ", "")
    sheet_name = f"{safe_name}_{target_year_month[1]}月"[:31]  # Excelシート名は31文字
    fname_base = f"明細表_{safe_name}_{target_year_month[0]}_{target_year_month[1]:02d}"

    xlsx_path = output_dir / f"{fname_base}.xlsx"
    pdf_path = output_dir / f"{fname_base}.pdf"

    work_days, log = populate_template(
        xlsx_path, sheet_name, dispatch, driver_name,
        target_year_month, prescribed_days
    )

    # 数式再計算
    recalc_excel(xlsx_path)

    # PDF変換
    convert_to_pdf(xlsx_path, pdf_path)

    # サマリ取得
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    summary = {
        "稼働日数": ws["L50"].value,
        "出勤加算ポイント": ws["N50"].value,
        "基本ポイント合計": ws["O50"].value,
        "追加ポイント合計": ws["S50"].value,
        "配車貢献ポイント": ws["U50"].value,
        "ポイント合計": ws["U53"].value,
    }
    low_conf = sum(1 for e in log if e["信頼度"] == "low")

    return {
        "xlsx_path": xlsx_path,
        "pdf_path": pdf_path,
        "driver_name": driver_name,
        "target_year_month": target_year_month,
        "work_days": work_days,
        "prescribed_days": prescribed_days,
        "summary": summary,
        "warnings": warnings,
        "low_conf_count": low_conf,
    }
